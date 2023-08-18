# Package Import
import sys # Included
from datetime import datetime # "pip install datetime" or https://pypi.org/project/datetime/
from pathlib import Path # "pip install pathlib" or https://pypi.org/project/pathlib/
from selenium import webdriver # "pip install selenium" or https://pypi.org/project/selenium/
from selenium.webdriver.common.by import By # For locators
from openpyxl import Workbook # "pip install openpyxl" or https://pypi.org/project/openpyxl/
from openpyxl import load_workbook # For loading workbooks



# Custom Methods
alphabet = "ZABCDEFGHIJKLMNOPQRSTUVWXY" # For Excel cell column indexing
def cell_write(x, y, t): # Method for writing to a specific cell (18278 * 1048576)
	out = "" # Initialize the column index
	if x > 702: # Case for three letters
		if (x - 26) % 676 == 0: # Case for Zs
			out += alphabet[((x - 26) // 676 - 1) % 26] # Set third letter
		else: # Case for other three letters
			out += alphabet[((x - 26) // 676) % 26] # Set third letter
	if x > 26: # Case for two letters
		if x % 26 == 0: # Case for Zs
			out += alphabet[(x // 26 - 1) % 26] # Set second letter
		else: # Case for other two letters
			out += alphabet[(x // 26) % 26] # Set second letter
	c = ws[out + alphabet[x % 26] + str(y)] # Access a specific cell
	c.value = t # Log input text

def file_write(file, line): # Method for writing to the log file
	f = open(file, "a") # Open the file
	f.write(line + "\n") # Write the line
	f.close() # Close the file



# Initialization Check
today = datetime.now().strftime("%b %d, %Y") # Get today's date as string
file_name = today + " - Log.txt" # Create the log file name
f = open(file_name, "w") # Open the file
f.write("Log for " + today + ":\n") # Write the line
f.close() # Close the file

if not Path("./Projects - Offshore Technology.xlsx").is_file(): # Case of no existing file
	print("Creating new file...") # Indicate new file creation
	file_write(file_name, "Creating new file...") # Sync to log file
	wb = Workbook() # Create a new workbook
	ws = wb.active # Access the default worksheet
	old_urls = [] # Empty the old urls list
	categories = ["S/N", "Project Name", "URL", "Date Published", "Project Abstract"] # Initialize the categories list
	for cat in categories: # Iterate over all categories
		cell_write(categories.index(cat) + 1, 1, cat) # Log default categories



# Update Check
else: # Case of file existing
	wb = load_workbook(filename = "Projects - Offshore Technology.xlsx") # Load existing workbook
	ws = wb[wb.sheetnames[0]] # Load the worksheet
	old_urls = [u.value for u in ws["C"]] # Build the old urls list
	categories = [cat.value for cat in ws[1]] # Build the categories list

ws.title = today # Set today as the sheetname
for i in range(1, len(old_urls) + 1): # Reindex
		cell_write(1, i + 1, i) # Index
new_urls = [] # Initialize the new urls list
positions = [] # Initialize the insert positions list

driver = webdriver.Chrome() # Initialize WebDriver on Chrome
# driver = webdriver.Edge() # Initialize WebDriver on Edge
# driver = webdriver.Firefox() # Initialize WebDriver on Firefox
# driver = webdriver.Ie() # Initialize WebDriver on Ie
# driver = webdriver.Safari() # Initialize WebDriver on Safari

driver.get("https://www.offshore-technology.com/projects-a-z/") # Navigate to the project catalog page
cur_urls = [figure.find_element(By.TAG_NAME, "a").get_attribute("href") for figure in driver.find_elements(By.TAG_NAME, "figure")] # Get current urls
for u in cur_urls: # Iterate over all current urls

	if u in old_urls: # Case of logged project
		continue # Move on
	else: # Case of new project
		new_urls.append(u) # Log the new url
		positions.append(cur_urls.index(u) + 2) # Log the insert position

if len(new_urls) == 0: # Case of no update
	file_write(file_name, today + ": No update to the projects!") # Sync to log file
	sys.exit(today + ": No update to the projects!") # Exit the code



# Project Logging
log = 0 # For reporting back work done
work = len(new_urls) # Get the workload
print(today + ": There are " + str(work) + " new projects") # Inform the workload
file_write(file_name, today + ": There are " + str(work) + " new projects") # Sync to log file

for url in new_urls: # Iterate through the new urls list
	driver.get(url) # Nevigate to a specific project page
	if driver.find_elements(By.TAG_NAME, "h1") == []: # Case of 404
		file_write(file_name, "This project page is currently 404: " + url + ", please try again later!") # Sync to log file
		sys.exit("This project page is currently 404: " + url + ", please try again later!") # Exit the code

	position = positions[new_urls.index(url)] # Get the insert position
	ws.insert_rows(position) # Insert new row
	cell_write(1, position, position - 1) # Index
	cell_write(2, position, driver.find_element(By.TAG_NAME, "h1").text) # Log the project name
	cell_write(3, position, url) # Log the url
	cell_write(4, position, driver.find_element(By.CLASS_NAME, "date-published").text) # Log the date published
	cell_write(5, position, driver.find_element(By.TAG_NAME, "p").text) # Log the project abstract

	if len(driver.find_elements(By.CLASS_NAME, "collapse-toggler-text")) != 0: # Check if expand is present
		driver.find_element(By.CLASS_NAME, "collapse-toggler-text").click() # Expand all boxes
	# driver.find_element(By.TAG_NAME, "main").find_element(By.TAG_NAME, "header").screenshot("./Screenshots/" + driver.title + ".png") # Screenshot
	for box in driver.find_elements(By.CLASS_NAME, "info-box"): # Iterate over all boxes
		category = box.find_element(By.TAG_NAME, "h5").text # Find the category of the box
		if category not in categories: # Case of new category
			categories.append(category) # Append the new category to the categories list
			cell_write(len(categories), 1, category) # Log the new category
		cell_write(categories.index(category) + 1, position, box.find_element(By.TAG_NAME, "p").text) # Log the content of the box

	wb.save("Projects - Offshore Technology.xlsx") # Save to file
	log += 1 # Increase log
	print("New project logged at #" + str(position - 1) + ", " + str(log) + " projects logged this session, " + str(work - log) + " remaining") # Report
	file_write(file_name, "New project logged at #" + str(position - 1) + ", " + str(log) + " projects logged this session, " + str(work - log) + " remaining") # Sync to log file
	
for i in range(1, len(cur_urls) + 1): # Reindex
		cell_write(1, i + 1, i) # Index
for d in range(len(new_urls)): # Iterate over extra lines
	ws.delete_rows(len(cur_urls) + 2) # Prune

wb.save("Projects - Offshore Technology.xlsx") # Save to file
print(today + ": Update finished, " + str(log) + " new projects updated, " + str(work - log) + " missed") # For checking missing log
file_write(file_name, today + ": Update finished, " + str(log) + " new projects updated, " + str(work - log) + " missed") # Sync to log file
print(today + ": There are " + str(len(cur_urls)) + " projects in total") # Final note
file_write(file_name, today + ": There are " + str(len(cur_urls)) + " projects in total") # Sync to log file
