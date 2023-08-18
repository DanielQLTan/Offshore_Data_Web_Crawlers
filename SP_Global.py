# Package Import
from selenium import webdriver# "pip install selenium" or https://pypi.org/project/selenium/

driver = webdriver.Chrome() # Initialize WebDriver on Chrome
# driver = webdriver.Edge() # Initialize WebDriver on Edge
# driver = webdriver.Firefox() # Initialize WebDriver on Firefox
# driver = webdriver.Ie() # Initialize WebDriver on Ie
# driver = webdriver.Safari() # Initialize WebDriver on Safari

from selenium.webdriver.common.by import By # For locators
from openpyxl import Workbook # "pip install openpyxl" or https://pypi.org/project/openpyxl/
wb = Workbook() # Create a new workbook



# Custom Methods
alphabet = "ZABCDEFGHIJKLMNOPQRSTUVWXY" # For Excel cell column indexing
def cell_write(x, y, t): # Method for writing to a specific cell (18278 * 1048576)
	out = "" # Initialize the column index
	if x > 702: # Case for three letters
		if (x - 26) % 676 == 0: # Case for Zs
			out += alphabet[((x - 26) // 676 - 1) % 26] # Set third letter
		else: # Case for other three letters
			out += alphabet[((x - 26) // 676) % 26] # Set third letter
	if x > 26: # Case for two letters
		if x % 26 == 0: # Case for Zs
			out += alphabet[(x // 26 - 1) % 26] # Set second letter
		else: # Case for other two letters
			out += alphabet[(x // 26) % 26] # Set second letter
	c = ws[out + alphabet[x % 26] + str(y)] # Access a specific cell
	c.value = t # Log corresponding text



# Table Logging
driver.get("https://www.spglobal.com/commodityinsights/en/ci/products/offshore-oil-rig-data.html") # Navigate to page
date = driver.find_element(By.CLASS_NAME, "product-intro").find_elements(By.TAG_NAME, "h4")[1].text # Find the date
names_text = [] # Initialize final worksheet names
name_num = 2 # Initialize name count
names = driver.find_elements(By.TAG_NAME, "h4") # Find all names
for name in names[name_num:]: # Iterate over all names

	wb.create_sheet(name.text) # Create the worksheets
	names_text.append(name.text) # Add to final worksheet names

tables = driver.find_elements(By.TAG_NAME, "table") # Find all tables
for table in tables: # Iterate over all tables

	ws = wb[names[name_num].text] # Get the worksheet
	heads = table.find_element(By.TAG_NAME, "tr").find_elements(By.TAG_NAME, "th") # Find all heads
	col_num = 1 # Set the col index
	for head in heads: # Iterate over all heads

		cell_write(col_num, 1, head.text) # Log the head
		col_num += 1 # Increase the col index

	rows = table.find_elements(By.TAG_NAME, "tr") # Find all rows
	for row in rows[1:]: # Iterate over all rows

		datas = row.find_elements(By.TAG_NAME, "td") # Find all datas
		col_num = 1 # Set the col index
		for data in datas: # Iterate over all datas

			cell_write(col_num, rows.index(row) + 1, data.text) # Log the data
			col_num += 1 # Increase the col index

	name_num += 1 # Increase name count

for s in wb.sheetnames: # Iterate over all worksheet names
	if s not in names_text: # Case if not in final worksheet names
		del wb[s] # Delete extra worksheets
if "Daniel Yergin, Ph.D." in wb.sheetnames: # Case of Daniel
	del wb["Daniel Yergin, Ph.D."] # Delete extra worksheets
wb.save(date + " Petrodata Weekly Rig Count.xlsx") # Save the Excel



# Image Logging
link = driver.find_element(By.LINK_TEXT, "View the Petrodata Offshore Rig Day Rate Trends Report").get_attribute("href") # Find link

main_image = driver.find_element(By.CSS_SELECTOR, "img[title=\"Offshore Rig Count\"]") # Find main image
main_name = main_image.get_attribute("alt") # Get main image name
driver.get(main_image.get_attribute("src")) # Navigate to main image
driver.find_element(By.TAG_NAME, "img").screenshot("./" + main_name + ".png") # Log main image

driver.get(link) # Navigate to link
month = driver.find_elements(By.TAG_NAME, "h2")[1].text # Find month
image_names = ["Worldwide Semisubmersibles", "Worldwide Drillships", "Southeast Asia Jackups", "Middle East Jackups"] # Image names

images = driver.find_elements(By.ID, "slick-slide") # Find all images
image_links = [image.find_element(By.TAG_NAME, "a").get_attribute("href") for image in images] # Create image links list
for l in image_links: # Iterate over all image links
	driver.get(l) # Navigate to image
	driver.find_element(By.TAG_NAME, "img").screenshot("./" + month + " " + image_names[image_links.index(l)] + ".png") # Log image
